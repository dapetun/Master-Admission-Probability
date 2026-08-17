"""Базовые тесты загрузки и симуляции."""

from __future__ import annotations

from pathlib import Path

import yaml

from admission_sim.load import (
    build_profiles,
    filter_profiles_by_status,
    load_dataset,
    program_and_campus_from_budget_xlsx,
    program_key,
    program_name_from_csv,
    _read_budget_xlsx,
)
from admission_sim.model import EXTERNAL, ApplicationRow
from admission_sim.scenarios import estimate_probability, what_if_consent
from admission_sim.simulate import (
    preference_list,
    resolve_focus_program,
    simulate_enrollment,
    subgraph_for_program,
)


def test_program_name_from_csv() -> None:
    path = Path("Математика_машинного обучения.2026-08-16_20-17-34.csv")
    assert program_name_from_csv(path) == "Математика_машинного обучения"


def _row(
    code: int,
    program: str,
    priority: int,
    score: float,
    rank: int,
    *,
    consent: bool = False,
    status: str = "Участвуете в конкурсе",
) -> ApplicationRow:
    return ApplicationRow(
        applicant_code=code,
        program=program,
        priority=priority,
        score=score,
        rank=rank,
        consent=consent,
        status=status,
    )


def test_missing_higher_priority_flag() -> None:
    rows = [
        _row(1, "A", priority=3, score=90, rank=1, consent=True),
        _row(2, "A", priority=1, score=80, rank=2, consent=True),
    ]
    profiles = build_profiles(rows)
    assert profiles[1].missing_higher_priority is True
    assert profiles[2].missing_higher_priority is False


def test_preference_list_inserts_external() -> None:
    rows = [_row(1, "A", priority=3, score=90, rank=1)]
    profile = build_profiles(rows)[1]
    # Один EXTERNAL: DA поглощает на первом разрыве, программа A недостижима.
    assert preference_list(profile) == [EXTERNAL]


def test_preference_list_large_priority_gap_stays_short() -> None:
    """Плейсхолдер приоритета 10**6 не раздувает список EXTERNAL."""
    rows = [_row(1, "A", priority=10**6, score=90, rank=1)]
    profile = build_profiles(rows)[1]
    prefs = preference_list(profile)
    assert prefs == [EXTERNAL]
    assert len(prefs) == 1


def test_cascade_simple_two_programs(tmp_path: Path) -> None:
    """
    A: 1 место, B: 1 место.

    code1: A pri1 score/rank лучше, согласие
    code2: A pri1 хуже, B pri2, согласие
    code3: B pri1, согласие

    Ожидание: 1→A, 3→B, 2 не проходит на A и идёт на B но место занято → None
    """
    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=True),
        _row(2, "B", 2, 95, 2, consent=True),
        _row(3, "B", 1, 80, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1}
    result = simulate_enrollment(profiles, seats, consent_only=True)
    assert result.destination(1) == "A"
    assert result.destination(3) == "B"
    assert result.destination(2) is None


def test_higher_score_bumps_after_fallback() -> None:
    """
    A: 1 место. B: 0 мест (внешняя дыра не нужна).

    X: pri1=B (нет мест / отвергнут), pri2=A, rank 1 на A
    Y: pri1=A, rank 2 на A

    X должен вытеснить Y на A.
    """
    rows = [
        _row(10, "B", 1, 50, 1, consent=True),
        _row(10, "A", 2, 100, 1, consent=True),
        _row(20, "A", 1, 80, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    result = simulate_enrollment(profiles, {"A": 1, "B": 0}, consent_only=True)
    assert result.destination(10) == "A"
    assert result.destination(20) is None


def test_external_absorbs_consent_with_gap() -> None:
    rows = [
        _row(1, "A", 3, 100, 1, consent=True),
        _row(2, "A", 1, 50, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    result = simulate_enrollment(profiles, {"A": 1}, consent_only=True)
    assert result.destination(1) == EXTERNAL
    assert result.destination(2) == "A"


def test_ovp_vs_vpp() -> None:
    rows = [
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1}
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    ovp = simulate_enrollment(profiles, seats, consent_only=False)
    assert vpp.destination(2) == "A"
    assert vpp.destination(1) is None  # нет в assignment активных без согласия
    assert 1 not in vpp.assignment or vpp.destination(1) is None
    assert ovp.destination(1) == "A"
    assert ovp.destination(2) is None


def test_what_if_consent() -> None:
    rows = [
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1}
    assert simulate_enrollment(profiles, seats).destination(2) == "A"
    forced = what_if_consent(profiles, seats, 1)
    assert forced.destination(1) == "A"
    assert forced.destination(2) is None


def test_monte_carlo_seed_stable() -> None:
    rows = [
        _row(1, "A", 1, 70, 3, consent=False),
        _row(2, "A", 1, 100, 1, consent=True),
        _row(3, "A", 1, 90, 2, consent=False),
        _row(4, "A", 1, 10, 4, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2}
    a = estimate_probability(profiles, seats, 1, n_simulations=50, seed=7)
    b = estimate_probability(profiles, seats, 1, n_simulations=50, seed=7)
    assert a.by_program == b.by_program
    assert a.any_loaded == b.any_loaded
    assert "Авто" in a.consent_model_description or "авто" in a.consent_model_description.lower()
    assert a.scenario == "auto"


def test_monte_carlo_workers_match_serial() -> None:
    """Один и тот же seed: serial (workers=1) и 2 процесса дают те же доли."""
    rows = [
        _row(1, "A", 1, 70, 3, consent=False),
        _row(2, "A", 1, 100, 1, consent=True),
        _row(3, "A", 1, 90, 2, consent=False),
        _row(4, "A", 1, 10, 4, consent=False),
        _row(5, "A", 1, 60, 5, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2}
    serial = estimate_probability(
        profiles, seats, 1, n_simulations=40, seed=11, n_workers=1
    )
    parallel = estimate_probability(
        profiles, seats, 1, n_simulations=40, seed=11, n_workers=2
    )
    assert serial.by_program == parallel.by_program
    assert serial.any_loaded == parallel.any_loaded
    assert serial.external == parallel.external
    assert serial.none == parallel.none


def test_monte_carlo_by_program_only_my_apps() -> None:
    """by_program не раздувается всеми ключами seats.yaml."""
    rows = [
        _row(1, "Mine", 1, 80, 2, consent=False),
        _row(2, "Mine", 1, 100, 1, consent=True),
        _row(3, "Other", 1, 90, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"Mine": 1, "Other": 5, "Unrelated": 10}
    est = estimate_probability(profiles, seats, 1, n_simulations=40, seed=1)
    assert list(est.by_program.keys()) == ["Mine"]
    assert "Other" not in est.by_program
    assert "Unrelated" not in est.by_program


def test_optimistic_lower_mean_than_pessimistic() -> None:
    from admission_sim.scenarios import estimate_consent_model

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=False),
        _row(3, "A", 1, 80, 3, consent=False),
        _row(4, "A", 1, 10, 4, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2}
    opt = estimate_consent_model(profiles, seats, scenario="optimistic")
    pes = estimate_consent_model(profiles, seats, scenario="pessimistic")
    assert opt.mean_undecided < pes.mean_undecided


def test_consent_model_higher_for_competitive() -> None:
    from admission_sim.scenarios import estimate_consent_model

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=True),
        _row(3, "A", 1, 80, 3, consent=False),
        _row(4, "A", 1, 10, 4, consent=False),
    ]
    profiles = build_profiles(rows)
    model = estimate_consent_model(profiles, {"A": 2})
    assert model.competitive_rate == 1.0
    assert model.noncompetitive_rate == 0.0
    assert model.by_code[3] == model.noncompetitive_rate
    assert model.by_code[4] == model.noncompetitive_rate


def test_auto_undecided_not_certain_when_mixed() -> None:
    """Смешанные согласия: auto < 1 для неопределившихся, pessimistic отличим."""
    from admission_sim.scenarios import estimate_consent_model

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=False),
        _row(3, "A", 1, 80, 3, consent=True),
        _row(4, "A", 1, 10, 4, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2}
    auto = estimate_consent_model(profiles, seats, scenario="auto")
    pes = estimate_consent_model(profiles, seats, scenario="pessimistic")
    opt = estimate_consent_model(profiles, seats, scenario="optimistic")
    bal = estimate_consent_model(profiles, seats, scenario="balanced")

    assert auto.by_code[2] < 1.0
    assert auto.by_code[4] < 1.0
    assert pes.by_code[2] > auto.by_code[2]
    assert pes.by_code[2] < 1.0
    assert pes.by_code[4] < 1.0
    assert auto.mean_undecided != pes.mean_undecided
    assert opt.by_code[2] < auto.by_code[2]
    assert 0.0 < bal.by_code[2] < 1.0
    assert bal.by_code[2] == bal.by_code[4]


def test_auto_does_not_copy_all_consent_onto_undecided() -> None:
    """p_comp=1.0 у победителей ОВП не означает p=1 у EXTERNAL / прочих."""
    from admission_sim.scenarios import estimate_consent_model

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=True),
        _row(3, "A", 2, 80, 3, consent=False),
        _row(4, "A", 1, 70, 4, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2}
    auto = estimate_consent_model(profiles, seats, scenario="auto")
    pes = estimate_consent_model(profiles, seats, scenario="pessimistic")
    assert auto.competitive_rate == 1.0
    assert auto.by_code[3] < 1.0
    assert auto.by_code[4] < 1.0
    assert pes.by_code[3] < 1.0
    assert pes.by_code[4] < 1.0
    assert pes.mean_undecided > auto.mean_undecided
    for code in (3, 4):
        assert pes.by_code[code] >= auto.by_code[code]


def test_load_seats_accepts_extended_entries(tmp_path: Path) -> None:
    from admission_sim.load import load_seats

    path = tmp_path / "seats.yaml"
    path.write_text(
        yaml.dump(
            {
                "programs": {
                    "A": 3,
                    "B": {"seats": 5, "title": "Test", "campus": "Москва"},
                }
            },
            allow_unicode=True,
        ),
        encoding="utf-8",
    )
    assert load_seats(path) == {"A": 3, "B": 5}


def test_load_seats_allows_null(tmp_path: Path) -> None:
    from admission_sim.load import load_seats

    path = tmp_path / "seats.yaml"
    path.write_text(
        yaml.dump({"programs": {"A": {"seats": None}}}, allow_unicode=True),
        encoding="utf-8",
    )
    assert load_seats(path) == {"A": None}


def test_unknown_seats_absorb_as_external() -> None:
    rows = [
        _row(1, "SPB", 1, 100, 1, consent=True),
        _row(1, "MSK", 2, 90, 2, consent=True),
        _row(2, "MSK", 1, 80, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    # SPB seats unknown → 1 goes EXTERNAL, 2 gets MSK
    result = simulate_enrollment(profiles, {"SPB": None, "MSK": 1}, consent_only=True)
    assert result.destination(1) == EXTERNAL
    assert result.destination(2) == "MSK"


def test_markdown_report_focuses_on_my_programs() -> None:
    from admission_sim.model import Dataset
    from admission_sim.report import MC_SCENARIO_UNCERTAINTY_NOTE, build_markdown_report
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(1, "Mine", 3, 100, 1, consent=False),
        _row(2, "Mine", 1, 90, 2, consent=True),
        _row(3, "Noise", 1, 80, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"Mine": 2, "Noise": 5, "Extra": 10}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["Mine", "Noise", "Extra"],
        source_files=["a.xlsx", "b.xlsx", "c.xlsx"],
        incomplete_priority_codes=[1],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    vpp_if = what_if_consent(profiles, seats, 1)
    ovp = simulate_enrollment(profiles, seats, consent_only=False)
    prob = estimate_probability(profiles, seats, 1, n_simulations=20, seed=3)
    threats = counterfactuals_for_threats(profiles, seats, 1, limit=5)
    md = build_markdown_report(
        dataset,
        1,
        vpp=vpp,
        vpp_if_consent=vpp_if,
        ovp=ovp,
        counterfactuals=threats,
        probability=prob,
        include_pending=False,
    )
    assert "| Mine |" in md
    assert "нет приоритетов: 1, 2" in md
    assert "### Ваши программы" in md
    assert "#### Mine" in md
    assert "свободно:" in md
    assert "1-м приоритетом и согласием: 0" in md
    assert "Из зачисленных сейчас ниже вас: 1" in md
    assert "Из зачисленных эту программу указали 1-м приоритетом" not in md
    assert "Ещё **2** конкурсов вне ваших программ" in md
    assert "Нет абитуриентов выше вас без согласия на ваших программах" in md
    # Не дампим чужие конкурсы как отдельные секции
    assert "#### Noise" not in md
    assert "#### Extra" not in md
    assert md.count("| Noise |") == 0
    assert "Источники: 3 файлов" in md
    assert MC_SCENARIO_UNCERTAINTY_NOTE in md


def test_compound_status_filter() -> None:
    """Составной статус MAGREPORTS не должен выпадать из модели."""
    rows = [
        _row(
            1,
            "A",
            1,
            45,
            1,
            status="Участвует в конкурсе / Ожидание результатов ВИ",
        ),
        _row(2, "A", 1, 40, 2, status="На рассмотрении"),
        _row(3, "A", 1, 0, 3, status="Ожидание результатов ВИ"),
    ]
    profiles = build_profiles(rows)
    active = filter_profiles_by_status(profiles, include_pending=False)
    assert 1 in active
    assert 2 not in active
    assert 3 not in active
    pending = filter_profiles_by_status(profiles, include_pending=True)
    assert 1 in pending
    assert 3 in pending
    assert 2 not in pending


def test_pending_apps_kept_by_default() -> None:
    """По умолчанию pending-заявки остаются в профиле (не «теряются»)."""
    rows = [
        _row(
            1,
            "Прикладные модели ИИ",
            1,
            0,
            10,
            status="Ожидание результатов ВИ",
        ),
        _row(1, "Финтех", 3, 100, 5, status="Участвует в конкурсе"),
    ]
    profiles = build_profiles(rows)
    kept = filter_profiles_by_status(profiles)
    assert 1 in kept
    programs = {a.program for a in kept[1].applications}
    assert programs == {"Прикладные модели ИИ", "Финтех"}
    dropped = filter_profiles_by_status(profiles, include_pending=False)
    assert {a.program for a in dropped[1].applications} == {"Финтех"}


def test_markdown_report_shows_pending_program() -> None:
    """Pending-программа видна в паспорте, MC и блоке мест при include_pending."""
    from admission_sim.model import Dataset
    from admission_sim.report import build_markdown_report
    from admission_sim.scenarios import counterfactuals_for_threats

    applied = "Прикладные модели искусственного интеллекта (Москва)"
    rows = [
        _row(1, applied, 1, 0, 115, status="Ожидание результатов ВИ"),
        _row(1, "ММО (Москва)", 2, 0, 50, status="Ожидание результатов ВИ"),
        _row(1, "Финтех (НН)", 3, 100, 5, status="Участвует в конкурсе"),
        _row(2, applied, 1, 90, 1, consent=True),
        _row(3, "Финтех (НН)", 1, 80, 2, consent=True),
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    seats = {applied: 5, "ММО (Москва)": 3, "Финтех (НН)": 15}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=[applied, "ММО (Москва)", "Финтех (НН)"],
        source_files=["a.xlsx", "b.xlsx", "c.xlsx"],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    vpp_if = what_if_consent(profiles, seats, 1)
    ovp = simulate_enrollment(profiles, seats, consent_only=False)
    prob = estimate_probability(profiles, seats, 1, n_simulations=20, seed=3)
    md = build_markdown_report(
        dataset,
        1,
        vpp=vpp,
        vpp_if_consent=vpp_if,
        ovp=ovp,
        counterfactuals=counterfactuals_for_threats(profiles, seats, 1, limit=5),
        probability=prob,
        include_pending=True,
    )
    assert applied in md
    assert "| Прикладные модели искусственного интеллекта (Москва) |" in md
    assert "#### Прикладные модели искусственного интеллекта (Москва)" in md
    assert "Учитывать ожидание результатов вступительных испытаний: да" in md
    assert "нет приоритетов:" not in md


def test_markdown_warns_when_pending_disabled() -> None:
    from admission_sim.model import Dataset
    from admission_sim.report import build_markdown_report

    rows = [_row(1, "Mine", 1, 100, 1)]
    profiles = build_profiles(rows)
    seats = {"Mine": 1}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["Mine"],
        source_files=["a.xlsx"],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    md = build_markdown_report(
        dataset,
        1,
        vpp=vpp,
        vpp_if_consent=vpp,
        ovp=vpp,
        counterfactuals=[],
        probability=None,
        include_pending=False,
    )
    assert "ожидание результатов вступительных испытаний" in md
    assert "**не учтены**" in md


def test_read_budget_xlsx_fixture(tmp_path: Path) -> None:
    """Мини-Budget.xlsx: шапка, ~id~, приоритет/сумма/согласие «Б»."""
    from openpyxl import Workbook

    path = tmp_path / "fixture_Budget.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = (
        'Список лиц, подавших документы\n'
        'Образовательная программа "Тестовая программа"\n'
        "Направление подготовки 01.04.02\n"
        "Москва"
    )
    headers = [
        "№ п/п",
        "Рег. номер",
        "Уникальный код поступающего",
        "Поступление на места по целевой квоте",
        "Приоритет бюджетного места",
        "Приоритет целевого места",
        "Конкурс портфолио",
        "Сумма конкурсных баллов",
        "Сумма конкурсных баллов в рамках квоты на целевые места",
        "Все оценки положительные",
        "Статус участия в конкурсе",
        "Согласие на зачисление",
    ]
    for col, name in enumerate(headers, start=1):
        ws.cell(7, col, name)
    ws.cell(8, 7, "~123~")
    data_rows = [
        (1, 100, 1111111, "-", 2, None, 100, 100, 100, "+", "Участвует в конкурсе", "Б"),
        (2, 101, 2222222, "-", 1, None, None, 0, 0, "-", "Ожидание результатов ВИ", None),
    ]
    for r_idx, values in enumerate(data_rows, start=9):
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx, value)
    wb.save(path)

    bare, campus = program_and_campus_from_budget_xlsx(path)
    assert bare == "Тестовая программа"
    assert campus == "Москва"
    rows = _read_budget_xlsx(path, program_key(bare, campus))
    assert len(rows) == 2
    assert rows[0].applicant_code == 1111111
    assert rows[0].priority == 2
    assert rows[0].score == 100.0
    assert rows[0].rank == 1
    assert rows[0].consent is True
    assert rows[0].status == "Участвует в конкурсе"
    assert rows[1].applicant_code == 2222222
    assert rows[1].priority == 1
    assert rows[1].score == 0.0
    assert rows[1].consent is False
    assert rows[1].status == "Ожидание результатов ВИ"


def test_load_dataset_from_fixtures(tmp_path: Path) -> None:
    csv_a = tmp_path / "ProgA.2026-08-16_12-00-00.csv"
    csv_a.write_text(
        "\n".join(
            [
                '"Порядковый номер";"Приоритет конкурса";"Подано согласие";'
                '"Сумма баллов";"Баллы за ВИ";"Баллы за ИД";"Статус";'
                '"Код поступающего";"Дата выбора конкурсной группы по Москве"',
                '1;1;"Электронное";100;"100";0;"Участвуете в конкурсе";111;"01.01.2026 в 12:00"',
                '2;1;"—";90;"90";0;"Участвуете в конкурсе";222;"01.01.2026 в 12:00"',
            ]
        ),
        encoding="utf-8",
    )
    seats_path = tmp_path / "seats.yaml"
    seats_path.write_text(
        yaml.dump({"programs": {"ProgA": 1}}, allow_unicode=True),
        encoding="utf-8",
    )
    dataset = load_dataset(tmp_path, seats_path)
    assert 111 in dataset.applicants
    assert dataset.applicants[111].consent is True
    assert dataset.seats["ProgA"] == 1
    result = simulate_enrollment(dataset.applicants, dataset.seats)
    assert result.destination(111) == "ProgA"


def test_program_fill_vacant_and_unknown_seats() -> None:
    from admission_sim.model import Dataset
    from admission_sim.report import format_seats, format_vacant, program_fill_stats

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=True),
        _row(3, "B", 1, 80, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 5, "B": None}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["A", "B"],
        source_files=[],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    a = program_fill_stats(dataset, vpp, "A", my_code=1)
    assert a.enrolled == 2
    assert a.vacant == 3
    assert a.priority1_ahead == 0
    assert a.enrolled_ahead == 0
    assert a.enrolled_below == 1
    assert format_seats(a.seats) == "5"
    b = program_fill_stats(dataset, vpp, "B", my_code=1)
    assert b.seats is None
    assert b.vacant is None
    assert format_seats(b.seats) == "число мест неизвестно"
    assert format_vacant(b.vacant) == "—"


def test_counterfactual_includes_rank_gap() -> None:
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 1}, 2, limit=5)
    assert len(threats) == 1
    item = threats[0]
    assert item.code == 1
    assert item.overlap_program == "A"
    assert item.their_rank == 1
    assert item.my_rank == 2
    assert item.gap == 1
    assert item.displaces_me is False


def test_threats_filter_skips_consent_and_non_overlap() -> None:
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 50, 3, consent=False),
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=False),
        _row(3, "B", 1, 80, 1, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 2, "B": 1}, 10, limit=30)
    assert [item.code for item in threats] == [2]


def test_threats_limit_keeps_largest_gaps() -> None:
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 10, 4, consent=False),
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=False),
        _row(3, "A", 1, 80, 3, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 1}, 10, limit=2)
    assert [item.code for item in threats] == [1, 2]


def test_threats_default_has_no_per_program_cap() -> None:
    from admission_sim.report import prepare_threats_view
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [_row(100, "A", 1, 10, 61, consent=False)]
    for i in range(1, 61):
        rows.append(_row(i, "A", 1, 90, i, consent=False))
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 80}, 100)
    assert len(threats) == 60
    assert {item.code for item in threats} == set(range(1, 61))
    view = prepare_threats_view(threats, None)
    assert len(view.shown) == 60
    assert {item.code for item in view.shown} == set(range(1, 61))
    assert view.count_caption is not None
    assert "60 из 60" in view.count_caption


def test_threats_person_on_two_programs_visible_when_filtering_either() -> None:
    from admission_sim.report import filter_counterfactuals_by_overlap
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 50, 10, consent=False),
        _row(10, "B", 2, 50, 100, consent=False),
        _row(1, "A", 2, 100, 5, consent=False),
        _row(1, "B", 1, 100, 1, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 20, "B": 20}, 10, limit=30)
    assert {(item.code, item.overlap_program) for item in threats} == {
        (1, "A"),
        (1, "B"),
    }
    filtered_a = filter_counterfactuals_by_overlap(threats, {"A"})
    assert len(filtered_a) == 1
    assert filtered_a[0].code == 1
    assert filtered_a[0].overlap_program == "A"
    assert filtered_a[0].their_rank == 5
    assert filtered_a[0].my_rank == 10
    assert filtered_a[0].destination is not None


def test_threats_per_program_keeps_rank_above_user_despite_global_gaps() -> None:
    from admission_sim.report import filter_counterfactuals_by_overlap
    from admission_sim.scenarios import counterfactuals_for_threats

    applied = "Прикладные модели искусственного интеллекта (Москва)"
    other = "Финтех (НН)"
    me = 100
    rows = [
        _row(me, applied, 1, 40, 93, consent=False),
        _row(me, other, 2, 40, 200, consent=False),
        _row(2, applied, 1, 100, 2, consent=False),
    ]
    for i in range(1, 41):
        rows.append(_row(1000 + i, other, 1, 90, i, consent=False))
    profiles = build_profiles(rows)
    seats = {applied: 20, other: 50}
    threats = counterfactuals_for_threats(profiles, seats, me, limit=30)
    filtered = filter_counterfactuals_by_overlap(threats, {applied})
    assert any(item.code == 2 for item in filtered)
    item = next(item for item in filtered if item.code == 2)
    assert item.their_rank == 2
    assert item.my_rank == 93
    assert item.overlap_program == applied
    assert item.destination is not None


def test_counterfactual_leaves_to_higher_priority_under_current_consents() -> None:
    from admission_sim.report import LEAVES_OVERLAP_LABEL, threat_if_consents_effect
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 50, 2, consent=True),
        _row(1, "B", 1, 100, 1, consent=False),
        _row(1, "A", 2, 100, 1, consent=False),
        _row(2, "B", 1, 40, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1}
    threats = counterfactuals_for_threats(profiles, seats, 10, limit=5)
    assert len(threats) == 1
    item = threats[0]
    assert item.code == 1
    assert item.overlap_program == "A"
    assert item.destination == "B"
    assert item.displaces_me is False
    assert threat_if_consents_effect(item) == LEAVES_OVERLAP_LABEL


def test_counterfactual_stays_on_overlap_if_other_program_blocked() -> None:
    from admission_sim.report import TAKES_OVERLAP_LABEL, threat_if_consents_effect
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 50, 2, consent=True),
        _row(1, "B", 1, 40, 2, consent=False),
        _row(1, "A", 2, 100, 1, consent=False),
        _row(2, "B", 1, 100, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1}
    threats = counterfactuals_for_threats(profiles, seats, 10, limit=5)
    assert len(threats) == 1
    item = threats[0]
    assert item.code == 1
    assert item.destination == "A"
    assert item.displaces_me is True
    assert threat_if_consents_effect(item) == TAKES_OVERLAP_LABEL


def test_threats_view_not_enrolled_hides_simulated_leaver() -> None:
    from admission_sim.report import prepare_threats_view
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 50, 3, consent=False),
        _row(1, "B", 1, 100, 1, consent=False),
        _row(1, "A", 2, 100, 1, consent=False),
        _row(2, "B", 1, 40, 2, consent=True),
        _row(3, "A", 1, 90, 2, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 2, "B": 1}
    threats = counterfactuals_for_threats(profiles, seats, 10)
    view = prepare_threats_view(threats, None)
    assert [item.code for item in view.shown] == [3]
    assert all(item.destination == item.overlap_program for item in view.shown)
    assert view.count_caption is not None
    assert "1 из 2" in view.count_caption


def test_threats_view_not_enrolled_shows_rank_columns() -> None:
    from admission_sim.report import prepare_threats_view, threat_table_rows
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 1}, 2, limit=5)
    view = prepare_threats_view(threats, None)
    assert view.empty_message is None
    assert view.caption is not None
    assert "выше вас без согласия" in view.caption
    rows_ui = threat_table_rows(view)
    assert len(rows_ui) == 1
    assert rows_ui[0]["Разрыв"] == 1
    assert "Согласие" not in rows_ui[0]
    assert rows_ui[0]["Если согласится"] == "займёт программу пересечения"
    assert "Вытесняет" not in rows_ui[0]
    assert "Вы до" not in rows_ui[0]
    assert "не весь список выше вас" in (view.caption or "")
    assert "только этот человек" in (view.caption or "")
    assert "ушёл бы на другую программу или вовне" in (view.caption or "")
    assert view.count_caption is not None
    assert "1 из 1" in view.count_caption
    assert "Случайные прогоны учитывают всех 1" in view.count_caption


def test_threats_view_enrolled_keeps_only_displacing() -> None:
    from admission_sim.report import prepare_threats_view, threat_table_rows
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(2, "A", 1, 90, 2, consent=True),
        _row(1, "A", 1, 100, 1, consent=False),
    ]
    profiles = build_profiles(rows)
    threats = counterfactuals_for_threats(profiles, {"A": 1}, 2, limit=5)
    view = prepare_threats_view(threats, "A")
    assert view.user_enrolled is True
    assert view.empty_message is None
    assert "вытесняют" in (view.caption or "")
    assert "не весь список выше вас" in (view.caption or "")
    assert all(item.displaces_me for item in view.shown)
    rows_ui = threat_table_rows(view)
    assert rows_ui[0]["Вытесняет"] == "да"
    assert rows_ui[0]["Если согласится"] == "займёт программу пересечения"


def test_threats_view_enrolled_hides_vacuous_rows() -> None:
    from admission_sim.report import prepare_threats_view
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(10, "A", 1, 80, 2, consent=True),
        _row(1, "A", 2, 100, 1, consent=False),
        _row(1, "B", 1, 100, 1, consent=False),
        _row(2, "B", 1, 50, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1}
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    assert vpp.destination(10) == "A"
    threats = counterfactuals_for_threats(profiles, seats, 10, limit=5)
    assert threats
    assert all(not item.displaces_me for item in threats)
    view = prepare_threats_view(threats, "A")
    assert view.shown == []
    assert view.empty_message is not None
    assert "никто не вытесняет" in view.empty_message


def test_threats_view_empty() -> None:
    from admission_sim.report import NO_RANK_THREATS_MESSAGE, prepare_threats_view

    view = prepare_threats_view([], None)
    assert view.shown == []
    assert view.empty_message == NO_RANK_THREATS_MESSAGE


def test_threats_view_filtered_empty_uses_distinct_message() -> None:
    from admission_sim.report import (
        FILTERED_EMPTY_THREATS_MESSAGE,
        NO_RANK_THREATS_MESSAGE,
        filter_counterfactuals_by_overlap,
        prepare_threats_view,
    )

    items = [_counterfactual(1, "B")]
    view = prepare_threats_view(
        filter_counterfactuals_by_overlap(items, {"A"}),
        None,
        filtered=True,
    )
    assert view.shown == []
    assert view.empty_message == FILTERED_EMPTY_THREATS_MESSAGE
    assert view.empty_message != NO_RANK_THREATS_MESSAGE
    assert prepare_threats_view([], None).empty_message == NO_RANK_THREATS_MESSAGE


_COUNTERFACTUAL_DEST_DEFAULT = object()


def _counterfactual(
    code: int,
    overlap_program: str,
    *,
    displaces_me: bool = False,
    destination: str | None | object = _COUNTERFACTUAL_DEST_DEFAULT,
):
    from admission_sim.scenarios import Counterfactual

    dest: str | None
    if destination is _COUNTERFACTUAL_DEST_DEFAULT:
        dest = overlap_program
    else:
        dest = destination  # type: ignore[assignment]
    return Counterfactual(
        code=code,
        destination=dest,
        displaces_me=displaces_me,
        my_destination_before="A" if displaces_me else None,
        my_destination_after=None if displaces_me else None,
        overlap_program=overlap_program,
        their_rank=1,
        my_rank=5,
        gap=4,
    )


def test_filter_counterfactuals_by_overlap() -> None:
    from admission_sim.report import filter_counterfactuals_by_overlap

    items = [
        _counterfactual(1, "A"),
        _counterfactual(2, "B"),
        _counterfactual(3, "A"),
    ]
    assert filter_counterfactuals_by_overlap(items, set()) == []
    filtered = filter_counterfactuals_by_overlap(items, {"A"})
    assert [item.code for item in filtered] == [1, 3]
    assert filter_counterfactuals_by_overlap(items, {"A", "B"}) == items


def test_threats_view_counts_follow_overlap_filter() -> None:
    from admission_sim.report import (
        filter_counterfactuals_by_overlap,
        prepare_threats_view,
    )

    items = [
        _counterfactual(1, "A", displaces_me=True),
        _counterfactual(2, "B", displaces_me=True),
        _counterfactual(3, "A", displaces_me=False),
    ]
    full = prepare_threats_view(items, "A")
    assert "2 из 3" in (full.caption or "")
    view = prepare_threats_view(
        filter_counterfactuals_by_overlap(items, {"A"}),
        "A",
    )
    assert "1 из 2" in (view.caption or "")
    assert [item.code for item in view.shown] == [1]
    assert all(item.overlap_program == "A" for item in view.shown)


def test_threats_view_not_enrolled_filter_keeps_matching_rows() -> None:
    from admission_sim.report import (
        filter_counterfactuals_by_overlap,
        prepare_threats_view,
    )

    items = [
        _counterfactual(1, "A"),
        _counterfactual(2, "B"),
        _counterfactual(3, "A"),
    ]
    view = prepare_threats_view(
        filter_counterfactuals_by_overlap(items, {"A"}),
        None,
    )
    assert [item.code for item in view.shown] == [1, 3]
    assert view.empty_message is None


def test_threats_view_not_enrolled_keeps_only_stay_on_overlap() -> None:
    from admission_sim.report import (
        filter_counterfactuals_staying_on_overlap,
        prepare_threats_view,
        stays_on_overlap_program,
        threats_stay_count_caption,
    )

    items = [
        _counterfactual(1, "A"),
        _counterfactual(2, "A", destination="B"),
        _counterfactual(3, "A", destination=EXTERNAL),
        _counterfactual(4, "A", destination=None),
    ]
    assert [item.code for item in filter_counterfactuals_staying_on_overlap(items)] == [1]
    assert stays_on_overlap_program(items[0]) is True
    assert stays_on_overlap_program(items[1]) is False
    view = prepare_threats_view(items, None)
    assert [item.code for item in view.shown] == [1]
    assert view.empty_message is None
    assert view.count_caption == threats_stay_count_caption(1, 4)
    assert "1 из 4" in (view.count_caption or "")
    assert "ушёл бы на другую программу или вовне" in (view.caption or "")


def test_threats_view_not_enrolled_all_leavers_has_distinct_message() -> None:
    from admission_sim.report import THREATS_ALL_LEAVE_MESSAGE, prepare_threats_view

    items = [
        _counterfactual(1, "A", destination="B"),
        _counterfactual(2, "A", destination=EXTERNAL),
    ]
    view = prepare_threats_view(items, None)
    assert view.shown == []
    assert view.empty_message is not None
    assert THREATS_ALL_LEAVE_MESSAGE in view.empty_message
    assert "0 из 2" in (view.count_caption or "")


def test_markdown_threats_not_enrolled_uses_rank_table() -> None:
    from admission_sim.model import Dataset
    from admission_sim.report import THREATS_CAPTION_NOT_ENROLLED, build_markdown_report
    from admission_sim.scenarios import counterfactuals_for_threats

    rows = [
        _row(1, "A", 1, 100, 1, consent=False),
        _row(2, "A", 1, 90, 2, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["A"],
        source_files=["a.xlsx"],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    md = build_markdown_report(
        dataset,
        2,
        vpp=vpp,
        vpp_if_consent=what_if_consent(profiles, seats, 2),
        ovp=simulate_enrollment(profiles, seats, consent_only=False),
        counterfactuals=counterfactuals_for_threats(profiles, seats, 2, limit=5),
        probability=None,
        include_pending=True,
    )
    assert THREATS_CAPTION_NOT_ENROLLED in md
    assert "1 из 1" in md
    assert "Случайные прогоны учитывают всех 1" in md
    assert "| Разрыв |" in md
    assert "| Если согласится |" in md
    assert "Вытесняет" not in md
    assert "число мест неизвестно" not in md
    assert "свободно: **1**" in md


def test_pending_keeps_file_score_no_extrapolation() -> None:
    """Pending участвует с баллами и местом из файла, без догадки о будущих баллах."""
    rows = [
        _row(1, "A", 1, 0, 10, status="Ожидание результатов ВИ"),
        _row(2, "A", 1, 90, 1, consent=True, status="Участвует в конкурсе"),
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    assert profiles[1].applications[0].score == 0.0
    assert profiles[1].applications[0].rank == 10
    result = simulate_enrollment(profiles, {"A": 1}, consent_only=False)
    assert result.destination(2) == "A"
    assert result.destination(1) is None


def test_priority2_hundred_leaves_when_wins_priority1() -> None:
    """100 баллов на приоритете 2 не занимает место, если проходит на приоритет 1."""
    rows = [
        _row(100, "A", 1, 100, 1, consent=True),
        _row(100, "B", 2, 100, 1, consent=True),
        _row(1, "B", 1, 80, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    result = simulate_enrollment(profiles, {"A": 1, "B": 1}, consent_only=True)
    assert result.destination(100) == "A"
    assert result.destination(1) == "B"


def test_priority2_hundred_takes_program_if_loses_priority1() -> None:
    """Если приоритет 1 не достался — 100 баллов занимают приоритет 2."""
    rows = [
        _row(50, "A", 1, 100, 1, consent=True),
        _row(100, "A", 1, 90, 2, consent=True),
        _row(100, "B", 2, 100, 1, consent=True),
        _row(1, "B", 1, 80, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    result = simulate_enrollment(profiles, {"A": 1, "B": 1}, consent_only=True)
    assert result.destination(50) == "A"
    assert result.destination(100) == "B"
    assert result.destination(1) is None


def test_worse_rank_cannot_displace_better_rank() -> None:
    """Худшее место не выталкивает лучшее, даже если у худшего 1-й приоритет."""
    rows = [
        _row(10, "B", 1, 50, 2, consent=True),
        _row(10, "A", 2, 100, 1, consent=True),
        _row(20, "A", 1, 40, 10, consent=True),
        _row(30, "B", 1, 90, 1, consent=True),
    ]
    profiles = build_profiles(rows)
    result = simulate_enrollment(profiles, {"A": 1, "B": 1}, consent_only=True)
    assert result.destination(30) == "B"
    assert result.destination(10) == "A"
    assert result.destination(20) is None

    leftover = simulate_enrollment(profiles, {"A": 2, "B": 1}, consent_only=True)
    assert leftover.destination(10) == "A"
    assert leftover.destination(20) == "A"
    assert leftover.destination(30) == "B"


def test_program_fill_counts_priority1_ahead_not_below() -> None:
    """1-й приоритет относительно пользователя: ниже по месту не входят."""
    from admission_sim.model import Dataset
    from admission_sim.report import program_fill_stats

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "B", 1, 80, 2, consent=True),
        _row(2, "A", 2, 90, 2, consent=True),
        _row(20, "B", 1, 95, 1, consent=True),
        _row(3, "A", 1, 85, 3, consent=False),
        _row(10, "C", 1, 70, 1, consent=True),
        _row(10, "A", 2, 70, 4, consent=True),
        _row(4, "A", 1, 11, 8, consent=True),
        _row(5, "A", 1, 3, 9, consent=True),
        _row(
            6,
            "A",
            1,
            0,
            10,
            consent=True,
            status="Ожидание результатов ВИ",
        ),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 6, "B": 1, "C": 0}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["A", "B", "C"],
        source_files=[],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    assert vpp.destination(1) == "A"
    assert vpp.destination(2) == "A"
    assert vpp.destination(20) == "B"
    assert vpp.destination(10) == "A"
    assert vpp.destination(4) == "A"
    assert vpp.destination(5) == "A"
    assert vpp.destination(6) == "A"

    stats = program_fill_stats(dataset, vpp, "A", my_code=10)
    assert stats.enrolled == 6
    assert stats.vacant == 0
    assert stats.priority1_ahead == 1
    assert stats.enrolled_ahead == 2
    assert stats.enrolled_below == 3


def test_focus_subgraph_matches_full_on_target() -> None:
    """Подграф для B совпадает с полной симуляцией по зачислению на B."""
    rows = [
        _row(1, "B", 1, 70, 3, consent=True),
        _row(100, "A", 1, 100, 1, consent=True),
        _row(100, "B", 2, 100, 1, consent=True),
        _row(2, "A", 2, 80, 2, consent=True),
        _row(2, "D", 1, 80, 1, consent=True),
        _row(3, "D", 1, 90, 2, consent=True),
        _row(9, "C", 1, 50, 1, consent=True),
        _row(10, "C", 1, 40, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1, "C": 2, "D": 1}
    full = simulate_enrollment(profiles, seats, consent_only=True)
    reduced_a, reduced_s = subgraph_for_program(profiles, seats, "B")
    assert "C" not in reduced_s
    assert "A" in reduced_s
    assert "D" in reduced_s
    assert 9 not in reduced_a
    red = simulate_enrollment(reduced_a, reduced_s, consent_only=True)
    for code in profiles:
        on_full = full.destination(code) == "B"
        on_red = red.destination(code) == "B"
        assert on_full == on_red, code
    assert full.destination(100) == "A"
    assert red.destination(100) == "A"
    assert full.destination(1) == "B"
    assert red.destination(1) == "B"
    full_ovp = simulate_enrollment(profiles, seats, consent_only=False)
    red_ovp = simulate_enrollment(reduced_a, reduced_s, consent_only=False)
    for code in profiles:
        assert (full_ovp.destination(code) == "B") == (red_ovp.destination(code) == "B")


def test_focus_subgraph_skips_unknown_seat_sink() -> None:
    """Чужих людей с программы без числа мест в подграф не тащим."""
    rows = [
        _row(1, "T", 2, 80, 2, consent=True),
        _row(1, "SINK", 1, 80, 1, consent=True),
        _row(2, "T", 1, 70, 3, consent=True),
        _row(99, "SINK", 1, 100, 1, consent=True),
        _row(99, "UNRELATED", 2, 100, 1, consent=True),
        _row(50, "UNRELATED", 1, 90, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"T": 1, "SINK": None, "UNRELATED": 1}
    reduced_a, reduced_s = subgraph_for_program(profiles, seats, "T")
    assert 99 not in reduced_a
    assert 50 not in reduced_a
    assert "UNRELATED" not in reduced_s
    assert "SINK" not in reduced_s
    full = simulate_enrollment(profiles, seats, consent_only=True)
    red = simulate_enrollment(reduced_a, reduced_s, consent_only=True)
    assert full.destination(1) == EXTERNAL
    assert red.destination(1) == EXTERNAL
    assert full.destination(2) == "T"
    assert red.destination(2) == "T"


def test_focus_subgraph_skips_zero_seat_only_applicants() -> None:
    """На конкурсе с 0 мест никто не держится — лишних людей не берём."""
    rows = [
        _row(100, "A", 1, 100, 1, consent=True),
        _row(100, "B", 2, 100, 1, consent=True),
        _row(1, "B", 1, 80, 2, consent=True),
        _row(9, "A", 1, 50, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 0, "B": 1}
    reduced_a, reduced_s = subgraph_for_program(profiles, seats, "B")
    assert 9 not in reduced_a
    assert "A" in reduced_s
    full = simulate_enrollment(profiles, seats, consent_only=True)
    red = simulate_enrollment(reduced_a, reduced_s, consent_only=True)
    assert full.destination(100) == "B"
    assert red.destination(100) == "B"
    assert full.destination(1) is None
    assert red.destination(1) is None


def test_focus_subgraph_keeps_external_gap() -> None:
    """Пробел приоритета в подграфе по-прежнему уводит во EXTERNAL."""
    rows = [
        _row(1, "B", 2, 100, 1, consent=True),
        _row(2, "B", 1, 80, 2, consent=True),
    ]
    profiles = build_profiles(rows)
    seats = {"B": 1}
    reduced_a, reduced_s = subgraph_for_program(profiles, seats, "B")
    full = simulate_enrollment(profiles, seats, consent_only=True)
    red = simulate_enrollment(reduced_a, reduced_s, consent_only=True)
    assert full.destination(1) == EXTERNAL
    assert red.destination(1) == EXTERNAL
    assert full.destination(2) == "B"
    assert red.destination(2) == "B"


def test_focus_program_mc_runs_and_filters() -> None:
    rows = [
        _row(1, "B", 2, 70, 2, consent=False),
        _row(1, "A", 1, 70, 2, consent=False),
        _row(2, "A", 1, 100, 1, consent=True),
        _row(3, "B", 1, 90, 1, consent=False),
        _row(9, "C", 1, 50, 1, consent=False),
    ]
    profiles = build_profiles(rows)
    seats = {"A": 1, "B": 1, "C": 5}
    est = estimate_probability(
        profiles,
        seats,
        1,
        n_simulations=40,
        seed=3,
        focus_program="B",
    )
    assert est.focus_program == "B"
    assert "B" in est.by_program
    assert "A" in est.by_program
    assert "C" not in est.by_program


def test_resolve_focus_program_substring() -> None:
    programs = [
        "Прикладные модели искусственного интеллекта (Москва)",
        "Финтех (Нижний Новгород)",
    ]
    assert (
        resolve_focus_program("Прикладные модели", programs)
        == programs[0]
    )
    assert resolve_focus_program(programs[1], programs) == programs[1]
    try:
        resolve_focus_program("нет такой", programs)
        raise AssertionError("ожидали KeyError")
    except KeyError:
        pass


def test_consent_rates_ignore_pending_zero_scores() -> None:
    """Pending с 0 не размывают доли и не считаются конкурентами из-за пустых мест."""
    from admission_sim.scenarios import (
        PESSIMISTIC_COMPETITIVE_FLOOR,
        UNDECIDED_PRIOR_OVERALL,
        estimate_consent_model,
    )

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=True),
        *[
            _row(10 + i, "A", 1, 0, 10 + i, status="Ожидание результатов ВИ")
            for i in range(8)
        ],
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    seats = {"A": 10}
    auto = estimate_consent_model(profiles, seats, scenario="auto")
    bal = estimate_consent_model(profiles, seats, scenario="balanced")
    pes = estimate_consent_model(profiles, seats, scenario="pessimistic")

    assert auto.scored_n == 2
    assert auto.pending_unknown_n == 8
    assert auto.competitive_rate == 1.0
    assert auto.overall_rate == 1.0
    assert auto.by_code[10] == UNDECIDED_PRIOR_OVERALL
    assert bal.by_code[10] == UNDECIDED_PRIOR_OVERALL
    assert pes.by_code[10] < 1.0
    assert pes.by_code[10] < PESSIMISTIC_COMPETITIVE_FLOOR
    assert "известными баллами" in auto.description
    assert "8" in auto.description


def test_pending_ahead_counts_above_user() -> None:
    from admission_sim.report import pending_ahead_rows

    rows = [
        _row(1, "A", 1, 80, 5),
        _row(2, "A", 1, 0, 2, status="Ожидание результатов ВИ"),
        _row(3, "A", 1, 0, 8, status="Ожидание результатов ВИ"),
        _row(4, "A", 1, 90, 1, consent=True),
        _row(1, "B", 2, 70, 3),
        _row(5, "B", 1, 0, 4, status="Ожидание результатов ВИ"),
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    stats = pending_ahead_rows(profiles, 1)
    by_program = {row.program: row for row in stats}
    assert by_program["A"].pending_ahead == 1
    assert by_program["A"].pending_on_program == 2
    assert by_program["B"].pending_ahead == 0
    assert by_program["B"].pending_on_program == 1


def test_markdown_pending_caption_and_ahead_column() -> None:
    from admission_sim.model import Dataset
    from admission_sim.report import PENDING_SCORE_CAPTION, build_markdown_report

    rows = [
        _row(1, "Mine", 1, 80, 5),
        _row(2, "Mine", 1, 0, 2, status="Ожидание результатов ВИ"),
        _row(3, "Mine", 1, 90, 1, consent=True),
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    seats = {"Mine": 2}
    dataset = Dataset(
        applicants=profiles,
        seats=seats,
        programs=["Mine"],
        source_files=["a.xlsx"],
        incomplete_priority_codes=[],
    )
    vpp = simulate_enrollment(profiles, seats, consent_only=True)
    md = build_markdown_report(
        dataset,
        1,
        vpp=vpp,
        vpp_if_consent=vpp,
        ovp=simulate_enrollment(profiles, seats, consent_only=False),
        counterfactuals=[],
        probability=None,
        include_pending=True,
    )
    assert PENDING_SCORE_CAPTION in md
    assert "Выше вас ждут экзамен" in md
    assert "| 1 из 1 |" in md


def test_pessimistic_not_certain_with_pending_majority() -> None:
    """Регрессия: пачка pending с 0 не должна давать пессимистичным p=1."""
    from admission_sim.scenarios import estimate_consent_model

    rows = [
        _row(1, "A", 1, 100, 1, consent=True),
        _row(2, "A", 1, 90, 2, consent=False),
        *[
            _row(20 + i, "A", 1, 0, 20 + i, status="Ожидание результатов ВИ")
            for i in range(20)
        ],
    ]
    profiles = filter_profiles_by_status(build_profiles(rows), include_pending=True)
    pes = estimate_consent_model(profiles, {"A": 5}, scenario="pessimistic")
    assert pes.by_code[2] < 1.0
    for i in range(20):
        assert pes.by_code[20 + i] < 1.0

